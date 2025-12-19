import json
import sys
from datetime import datetime, date, time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class DateTimeEncoder(json.JSONEncoder):
    """Custom JSON encoder for datetime objects."""
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif isinstance(obj, date):
            return obj.isoformat()
        elif isinstance(obj, time):
            return obj.isoformat()
        return super().default(obj)


def rgb_to_hex(rgb):
    """Convert RGB tuple or ARGB string to hex color string."""
    if rgb is None:
        return None
    if isinstance(rgb, str):
        # ARGB format (8 characters) - skip first 2 (alpha)
        if len(rgb) == 8:
            return f"#{rgb[2:]}"
        elif len(rgb) == 6:
            return f"#{rgb}"
        return rgb
    return None


def get_cell_background_color(cell):
    """Extract background color from cell."""
    try:
        fill = cell.fill
        if fill and fill.fgColor:
            if fill.fgColor.type == "rgb" and fill.fgColor.rgb:
                rgb = fill.fgColor.rgb
                if rgb != "00000000":  # Not transparent
                    return rgb_to_hex(rgb)
            elif fill.fgColor.type == "indexed":
                return f"indexed:{fill.fgColor.indexed}"
            elif fill.fgColor.type == "theme":
                return f"theme:{fill.fgColor.theme}"
    except Exception:
        pass
    return None


def get_cell_font_color(cell):
    """Extract font color from cell."""
    try:
        font = cell.font
        if font and font.color:
            if font.color.type == "rgb" and font.color.rgb:
                return rgb_to_hex(font.color.rgb)
            elif font.color.type == "indexed":
                return f"indexed:{font.color.indexed}"
            elif font.color.type == "theme":
                return f"theme:{font.color.theme}"
    except Exception:
        pass
    return None


def get_cell_font_size(cell):
    """Extract font size from cell."""
    try:
        if cell.font and cell.font.size:
            return cell.font.size
    except Exception:
        pass
    return None


def excel_to_json(excel_path, output_path=None):
    """Convert Excel file to JSON with cell styling information."""
    workbook = load_workbook(excel_path, data_only=False)
    result = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = {
            "sheet_name": sheet_name,
            "cells": []
        }

        for row in sheet.iter_rows():
            for cell in row:
                # Skip empty cells
                if cell.value is None and get_cell_background_color(cell) is None:
                    continue

                cell_address = f"{get_column_letter(cell.column)}{cell.row}"
                cell_info = {
                    "cell_address": cell_address,
                    "value": cell.value if cell.value is not None else "",
                    "background_color": get_cell_background_color(cell),
                    "font_size": get_cell_font_size(cell),
                    "font_color": get_cell_font_color(cell)
                }
                sheet_data["cells"].append(cell_info)

        result.append(sheet_data)

    workbook.close()

    # Output JSON
    json_output = json.dumps(result, ensure_ascii=False, indent=2, cls=DateTimeEncoder)

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(json_output)
        print(f"JSON saved to: {output_path}")
    else:
        print(json_output)

    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        # Default path
        excel_path = r"LegacyCode\NPL 평가 Report 요건정의.xlsx"
    else:
        excel_path = sys.argv[1]

    output_path = sys.argv[2] if len(sys.argv) > 2 else "output.json"

    excel_to_json(excel_path, output_path)
